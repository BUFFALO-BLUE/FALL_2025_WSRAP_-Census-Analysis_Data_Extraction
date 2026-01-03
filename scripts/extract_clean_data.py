import pandas as pd
import os

print("=" * 70)
print("📊 EXTRACTING CLEAN DATA FROM EXCEL USING YOUR MAPPING")
print("=" * 70)

# Your mapping information
HEADER_ROW = 4  # Row with headers (Row 4)
DATA_START_ROW = 5  # First data row (Row 5)
EXCEL_PATH = r"data\from_jeremy\transcriptions\Research Assistant Real Estate  (version 1).xlsx"

# Column mapping from your description
COLUMN_MAPPING = {
    'Race': 'B',
    'House_Number': 'C', 
    'Street_Name': 'D',
    'Owned_Home_Value': 'E',
    'Rented': 'F',
    'Notes': 'G'
}

# Try to extract clean data
try:
    print(f"📁 Reading: {os.path.basename(EXCEL_PATH)}")
    print(f"📐 Using your mapping:")
    for col_name, col_letter in COLUMN_MAPPING.items():
        print(f"   • {col_name}: Column {col_letter}")
    
    # Method 1: Read with column ranges
    # We'll skip rows 0-3 (0-indexed in pandas)
    df = pd.read_excel(
        EXCEL_PATH,
        skiprows=HEADER_ROW - 1,  # Skip to header row
        usecols=list(COLUMN_MAPPING.values()),  # Use only our columns
        names=list(COLUMN_MAPPING.keys()),  # Use proper column names
        engine='openpyxl'
    )
    
    print(f"\n✅ Successfully extracted {len(df)} rows!")
    print(f"📊 Data preview:")
    print(df.head(10).to_string())
    
    # Basic statistics
    print(f"\n📈 Data Summary:")
    for column in df.columns:
        non_null = df[column].count()
        unique = df[column].nunique()
        sample = df[column].dropna().unique()[:3]
        print(f"   {column}:")
        print(f"     • {non_null} non-null values")
        print(f"     • {unique} unique values")
        if len(sample) > 0:
            print(f"     • Sample: {sample}")
    
    # Save cleaned data
    output_csv = r"data\from_jeremy\transcriptions\clean_census_data.csv"
    df.to_csv(output_csv, index=False)
    print(f"\n💾 Clean data saved to: {output_csv}")
    
except Exception as e:
    print(f"❌ Error: {e}")
    print("\nTrying alternative method...")
    
    # Alternative: Manual extraction
    import openpyxl
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    # Extract data manually
    data = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        row_data = {}
        for col_name, col_letter in COLUMN_MAPPING.items():
            cell = ws[f"{col_letter}{row}"]
            row_data[col_name] = cell.value
        # Only add row if it has some data
        if any(v for v in row_data.values() if v):
            data.append(row_data)
    
    df = pd.DataFrame(data)
    print(f"✅ Manually extracted {len(df)} rows")
    print(df.head(10).to_string())
    
    # Save
    output_csv = r"data\from_jeremy\transcriptions\clean_census_data.csv"
    df.to_csv(output_csv, index=False)
    print(f"\n💾 Clean data saved to: {output_csv}")

print("\n" + "=" * 70)
print("🚨 CRITICAL PROBLEM IDENTIFIED")
print("=" * 70)

print("""
❌ MISSING CONNECTION: NO IMAGE FILENAMES

The Excel has transcriptions BUT lacks image references.
This means we DON'T KNOW which transcription goes with which census image.

EXAMPLE OF WHAT WE NEED:
| Image_Name           | Race  | House_Number | Street_Name      |
|----------------------|-------|--------------|------------------|
| m-t0627-00538-00634  | White | 131          | Chesire Street   |
| m-t0627-00538-00635  | Black | 108          | Ansonia          |

WHAT WE ACTUALLY HAVE:
| Race  | House_Number | Street_Name      |
|-------|--------------|------------------|
| White | 131          | Chesire Street   |
| Black | 108          | Ansonia          |
| White | 517          | New Britain Ave  |

✅ WHAT WE CAN DO IMMEDIATELY:
""")

# Create next steps
print("🎯 NEXT STEPS (Choose One):")

print("""
OPTION 1: ASK JEREMY FOR COMPLETE DATA (Recommended)
------------------------------------------
Email Jeremy with this specific request:

"Hi Jeremy,

I've extracted the census data from the Excel file, but it's missing 
the crucial column linking transcriptions to specific census images.

Could you please provide:
1. The COMPLETE Excel file with an 'Image_Name' column 
   (e.g., 'm-t0627-00538-00634.jpg')
2. Or a mapping file showing which Excel rows correspond to which 
   census image files

Without this mapping, I cannot connect the transcriptions to the 
extracted cell images for model training.

Thanks,
Musarah"

OPTION 2: MANUAL MAPPING (If Jeremy can't provide)
------------------------------------------
If Jeremy can't provide image mappings, we can:
1. Take a SUBSET of census images (e.g., 100 images)
2. Manually transcribe them using your Excel as reference
3. Create a small training dataset
4. Train initial model, then use it to help transcribe more

OPTION 3: CREATE MAPPING FROM FOLDER NAMES
------------------------------------------
If the Excel rows are IN THE SAME ORDER as your image folders:
1. List your extracted image folders in order
2. Assume Excel Row 1 = First folder, Row 2 = Second folder, etc.
3. Create mapping automatically

Let's check if this might work:
""")

# Check extracted folders
extracted_dir = r"data\extracted_cells"
if os.path.exists(extracted_dir):
    folders = sorted([f for f in os.listdir(extracted_dir) 
                     if os.path.isdir(os.path.join(extracted_dir, f))])
    print(f"   You have {len(folders)} extracted image folders")
    print(f"   Excel has {len(df)} rows of data")
    
    if len(folders) == len(df):
        print("   ⚠️  WARNING: Same count! They MIGHT be in same order.")
        print("   This is risky but could work if Jeremy confirms the order.")
    else:
        print("   ❌ Different counts - can't auto-map by order.")

print("""
OPTION 4: SEMI-AUTOMATED APPROACH
------------------------------------------
1. Use the clean CSV we just created
2. For each census image, show the user possible transcriptions
3. Let user select the correct one
4. Build mapping interactively

🎯 RECOMMENDED ACTION PLAN:
""")

# Create action plan
action_plan = f"""
1. IMMEDIATE (Today):
   • Run this script to confirm extraction works
   • Email Jeremy requesting image mappings
   
2. SHORT-TERM (1-2 days):
   • If Jeremy responds: Integrate mapping and create dataset
   • If no response: Start manual mapping of 100 images
   
3. MEDIUM-TERM (3-5 days):
   • Create training dataset with whatever mapping we have
   • Begin HPC model training
   
4. CONTINGENCY:
   • We have {len(df)} clean transcriptions
   • We have ~{len(folders) if 'folders' in locals() else '?'} image folders
   • Even partial mapping gives us training data
"""

print(action_plan)

print("\n" + "=" * 70)
print("⚡ LET'S RUN THE EXTRACTION FIRST")
print("=" * 70)

print("The script above will extract clean CSV from the Excel.")
print("Run it with: python scripts/extract_clean_data.py")
print("\nThen we'll decide our next move based on the results.")