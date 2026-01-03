import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def force_read_formatted_excel():
    """Force extraction from formatted Excel file"""
    
    excel_path = r"data\from_jeremy\transcriptions\Research Assistant Real Estate  (version 1).xlsx"
    
    print("=" * 70)
    print("🔄 FORCING EXTRACTION FROM FORMATTED EXCEL FILE")
    print("=" * 70)
    
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        return None
    
    # Load with openpyxl to get raw values
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active
    
    print(f"Active sheet: {ws.title}")
    print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
    
    # Strategy 1: Try to find data by looking for patterns
    print("\n🔍 Scanning for data patterns...")
    
    # Collect ALL non-empty cells with their positions
    all_data = []
    for row in ws.iter_rows(min_row=1, max_row=min(200, ws.max_row)):
        for cell in row:
            if cell.value and str(cell.value).strip():
                all_data.append({
                    'row': cell.row,
                    'col': cell.column,
                    'col_letter': cell.column_letter,
                    'value': str(cell.value).strip(),
                    'is_header': cell.row < 10  # Assume first 10 rows might be headers
                })
    
    print(f"Found {len(all_data)} non-empty cells in first 200 rows")
    
    # Group by column to see what's where
    columns_data = {}
    for item in all_data:
        col = item['col_letter']
        if col not in columns_data:
            columns_data[col] = []
        columns_data[col].append(item['value'])
    
    print("\n📊 Data by column (first 3 values per column):")
    for col in sorted(columns_data.keys()):
        values = columns_data[col]
        print(f"  Column {col}: {values[:3]}")
        if len(values) > 3:
            print(f"      ... and {len(values)-3} more")
    
    # Strategy 2: Try to extract structured data
    print("\n🧱 Attempting to extract structured rows...")
    
    # Find potential data rows (rows with multiple non-empty cells)
    potential_data_rows = {}
    for row_num in range(1, min(100, ws.max_row + 1)):
        row_values = []
        for col in range(1, min(15, ws.max_column + 1)):  # Check first 15 columns
            cell = ws.cell(row=row_num, column=col)
            if cell.value and str(cell.value).strip():
                row_values.append(str(cell.value).strip())
        
        if len(row_values) >= 3:  # Rows with at least 3 values might be data
            potential_data_rows[row_num] = row_values
    
    print(f"Found {len(potential_data_rows)} potential data rows")
    
    # Show sample
    if potential_data_rows:
        print("\nSample data rows:")
        for i, (row_num, values) in enumerate(list(potential_data_rows.items())[:5]):
            print(f"  Row {row_num}: {values}")
    
    # Strategy 3: Try to guess column headers
    print("\n🏷️  Looking for column headers...")
    
    header_candidates = []
    for row_num in range(1, 10):  # Check first 10 rows for headers
        for col in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row=row_num, column=col)
            if cell.value:
                cell_text = str(cell.value).strip().lower()
                # Check if this looks like a header
                header_keywords = ['race', 'house', 'street', 'own', 'rent', 'note', 'image', 'name', 'number']
                for keyword in header_keywords:
                    if keyword in cell_text:
                        header_candidates.append({
                            'row': row_num,
                            'col': col,
                            'col_letter': cell.column_letter,
                            'value': str(cell.value).strip(),
                            'keyword': keyword
                        })
    
    if header_candidates:
        print("Potential headers found:")
        for h in header_candidates:
            print(f"  {h['col_letter']}{h['row']}: '{h['value']}' (contains '{h['keyword']}')")
    else:
        print("No obvious headers found in first 10 rows")
    
    # Save raw extraction for manual inspection
    print("\n💾 Saving raw extraction for manual inspection...")
    
    # Create a simple CSV with all non-empty cells
    import csv
    
    raw_data_path = r"data\from_jeremy\transcriptions\raw_extraction.csv"
    with open(raw_data_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Row', 'Column', 'Value'])
        for item in all_data:
            writer.writerow([item['row'], item['col_letter'], item['value']])
    
    print(f"✅ Raw data saved to: {raw_data_path}")
    print(f"   Total cells extracted: {len(all_data)}")
    
    # Create a summary
    summary_path = r"data\from_jeremy\transcriptions\file_summary.txt"
    with open(summary_path, 'w') as f:
        f.write(f"Excel File Analysis Summary\n")
        f.write(f"File: {excel_path}\n")
        f.write(f"Total rows in sheet: {ws.max_row}\n")
        f.write(f"Total columns in sheet: {ws.max_column}\n")
        f.write(f"Non-empty cells found: {len(all_data)}\n\n")
        
        f.write("Data by column:\n")
        for col in sorted(columns_data.keys()):
            values = columns_data[col]
            f.write(f"  Column {col}: {len(values)} values\n")
            f.write(f"    Sample: {values[:5]}\n")
        
        f.write("\nPotential data rows:\n")
        for row_num, values in list(potential_data_rows.items())[:10]:
            f.write(f"  Row {row_num}: {values}\n")
    
    print(f"✅ Summary saved to: {summary_path}")
    
    # Strategy 4: Try pandas with various parameters
    print("\n🐼 Trying pandas with different parameters...")
    
    results = []
    
    # Try different skiprow values
    for skiprows in [0, 1, 2, 3, 4, 5]:
        try:
            df = pd.read_excel(excel_path, skiprows=skiprows, nrows=20, engine='openpyxl')
            non_null_count = df.count().sum()
            results.append({
                'skiprows': skiprows,
                'shape': df.shape,
                'non_null': non_null_count,
                'columns': list(df.columns)[:5] if len(df.columns) > 0 else []
            })
        except Exception as e:
            pass
    
    if results:
        print("Pandas reading attempts:")
        for r in results:
            print(f"  skiprows={r['skiprows']}: {r['shape']} shape, {r['non_null']} non-null cells")
            if r['columns']:
                print(f"    Columns: {r['columns']}")
    
    return all_data

def create_clean_template():
    """Create a clean CSV template for manual data entry if needed"""
    
    template = """image_name,row_number,race,house_number,street_name,owned_value,rented,notes
# Example: m-t0627-00538-00634.jpg,0,White,123,Main Street,5000,20,""
# Example: m-t0627-00538-00634.jpg,1,Black,123,Main Street,,15,"Child laborer"
# Fill in your data below:
# Use one row per census form row (0-39 rows per image)
# Leave empty if no data

"""
    
    template_path = r"data\from_jeremy\transcriptions\data_template.csv"
    with open(template_path, 'w') as f:
        f.write(template)
    
    print(f"\n📋 Created clean template: {template_path}")
    print("   You could manually transcribe a small subset here for initial testing")

if __name__ == "__main__":
    # Run the forced extraction
    data = force_read_formatted_excel()
    
    # Create template
    create_clean_template()
    
    print("\n" + "=" * 70)
    print("🎯 RECOMMENDED ACTION PLAN")
    print("=" * 70)
    
    print("""
🚨 PRIMARY RECOMMENDATION:
   SEND THE EMAIL TO JEREMY NOW requesting clean data.

⚡ WHY THIS IS CRITICAL:
   1. Color-coded/formatted Excel → Hard for automation
   2. WPS Office limitations → Can't edit/save clean version
   3. Time efficiency → Jeremy can provide clean data faster than we can reverse-engineer

🔧 TEMPORARY WORKAROUNDS (while waiting):
   
   OPTION A: Manual extraction from raw data
   1. Open: data/from_jeremy/transcriptions/raw_extraction.csv
   2. Look for patterns in the data
   3. Identify which columns contain which information
   
   OPTION B: Manual transcription of small subset
   1. Use: data/from_jeremy/transcriptions/data_template.csv
   2. Manually transcribe 50-100 rows from the Excel file
   3. Use this small dataset for initial model testing
   
   OPTION C: OCR without transcriptions (unsupervised)
   1. We could train on extracted images only
   2. Use clustering to find patterns
   3. Later add labels when available

📅 PROJECT TIMELINE SUGGESTION:
   
   TODAY:
   • Email Jeremy for clean data
   • Examine raw_extraction.csv to understand data structure
   
   TOMORROW (if no response):
   • Manually transcribe 100 rows using the template
   • Create small training dataset for initial testing
   
   DAY 3:
   • Begin model training with whatever data we have
   • Continue following up with Jeremy

💡 KEY INSIGHT:
   Your OCR pipeline (extracting cells) is the HARD part - that's working!
   Getting clean labels is the EASY part - just needs proper formatting.

📬 SEND THE EMAIL NOW, then:
   Run: python scripts/hpc_preparation.py
   (This prepares everything else while we wait for data)
""")